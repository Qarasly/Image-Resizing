import streamlit as st
import pandas as pd
from PIL import Image, ImageOps, ImageStat
import requests
from io import BytesIO
import re
import zipfile
import base64
from rembg import remove
import json
import openpyxl
import time
from google import genai

# --- 1. SECURE CONFIGURATION ---
IMGBB_API_KEY = st.secrets.get("IMGBB_API_KEY")
GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY")

if not IMGBB_API_KEY:
    st.error("🚨 **ImgBB API Key Missing!** Please add `IMGBB_API_KEY` to your Streamlit Secrets.")
    st.stop()

# --- HELPER FUNCTIONS ---
def get_direct_url(url):
    if pd.isna(url): return url
    url = str(url).strip().replace('\n', '').replace('\r', '').replace('%0A', '')
    if "drive.google.com" in url:
        match = re.search(r"/d/([a-zA-Z0-9_-]{25,})", url)
        if match:
            return f"https://drive.google.com/uc?export=download&id={match.group(1)}"
    return url

def detect_background_color(img):
    if img.mode != 'RGB': img = img.convert('RGB')
    w, h = img.size
    corners = [img.crop((0,0,5,5)), img.crop((w-5,0,w,5)), img.crop((0,h-5,5,h)), img.crop((w-5,h-5,w,h))]
    r, g, b = 0, 0, 0
    for corner in corners:
        stat = ImageStat.Stat(corner)
        r += stat.mean[0]; g += stat.mean[1]; b += stat.mean[2]
    return (int(r/4), int(g/4), int(b/4))

def hex_to_rgb(hex_str):
    hex_str = hex_str.lstrip('#')
    return tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))

def process_image_pipeline(img, tw, th, final_color_rgb, bg_mode):
    if bg_mode: img = remove(img)
    curr_color = detect_background_color(img) if final_color_rgb == "AUTO" else final_color_rgb
    img.thumbnail((tw, th), Image.Resampling.LANCZOS)
    new_img = Image.new("RGB", (tw, th), curr_color)
    paste_pos = ((tw - img.size[0]) // 2, (th - img.size[1]) // 2)
    
    if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
        img = img.convert("RGBA")
        new_img.paste(img, paste_pos, img)
    else:
        new_img.paste(img, paste_pos)
    return new_img

def upload_to_imgbb(img_buffer, psku):
    url = "https://api.imgbb.com/1/upload"
    payload = {
        "key": IMGBB_API_KEY,
        "image": base64.b64encode(img_buffer.getvalue()).decode("utf-8"),
        "name": f"sku_{psku}"
    }
    res = requests.post(url, data=payload)
    res.raise_for_status()
    return res.json()["data"]["url"]

def generate_dynamic_content(img, selected_fields, mapping_context=""):
    """Sends the image, fields, and mapping data to Gemini with an auto-retry loop for 503 errors."""
    if not GEMINI_API_KEY:
        return {}
        
    client = genai.Client(api_key=GEMINI_API_KEY)
    fields_schema = {field: f"Value generated for {field}" for field in selected_fields}
    
    prompt = f"""
    You are an e-commerce catalog AI. Analyze this product image.
    
    Here is the global background mapping rules/context for this project:
    {mapping_context}
    
    Based on the image and context, fill out the following requested fields and return them as a valid JSON object matching this exact structure:
    {json.dumps(fields_schema, indent=2)}
    
    Provide accurate information. Return ONLY the raw JSON object. Do not wrap it in markdown code blocks or triple backticks.
    """
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[img, prompt]
            )
            raw = response.text.strip()
            
            backticks = chr(96) * 3
            if raw.startswith(backticks):
                raw = re.sub(r"^" + re.escape(backticks) + r"(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*" + re.escape(backticks) + r"$", "", raw)
                
            return json.loads(raw.strip())
            
        except Exception as e:
            if "503" in str(e) or "UNAVAILABLE" in str(e):
                if attempt < max_retries - 1:
                    time.sleep(2 + attempt * 2)
                    continue
            return {field: f"AI Error: {str(e)}" for field in selected_fields}
            
    return {field: "AI Error: Server busy after multiple retries" for field in selected_fields}

@st.cache_data(show_spinner=False)
def process_url_full(src_url, psku, tw, th, final_color_rgb, bg_mode):
    try:
        resp = requests.get(get_direct_url(src_url), timeout=15)
        resp.raise_for_status()
        raw_img = Image.open(BytesIO(resp.content))
        proc_img = process_image_pipeline(raw_img, tw, th, final_color_rgb, bg_mode)
        
        buf = BytesIO(); proc_img.save(buf, format="JPEG", quality=90); buf.seek(0)
        link = upload_to_imgbb(buf, psku)
        return proc_img, link
    except Exception as e:
        return None, f"Error: {str(e)}"

# --- 2. UI LAYOUT ---
st.set_page_config(page_title="Bulk Resizing & Dynamic AI", layout="wide")
st.title("🏭 Bulk Image Studio & Template-Driven AI")

# STEP 1: GLOBAL CONFIG
with st.container(border=True):
    st.subheader("⚙️ Step 1: Mode & Image Configuration")
    s1, s2, s3 = st.columns([3, 4, 4])
    
    with s1:
        st.write("**Task Features**")
        use_resizer = st.checkbox("Run Image Resizer & Padding", value=True)
        use_ai_content = st.checkbox("Run Gemini AI Content Generation", value=False)
        
        if not use_resizer and not use_ai_content:
            st.warning("⚠️ Please choose at least one option above to continue.")
            
    with s2:
        st.write("**Processing Options**")
        input_mode = st.selectbox("Input Source", ["Links (Excel/CSV Sheet)", "Local Image Files"])
        output_mode = st.selectbox("Output Format", ["Links (Excel Sheet)", "Images (ZIP File)"])
        bg_mode = st.toggle("AI Background Removal (rembg)", value=False, disabled=not use_resizer, help="Isolates the product image.")
        
    with s3:
        st.write("**Canvas Background Fill**")
        color_type = st.radio("Style:", ["Standard White", "Automatic (Match Image)", "Custom Color"], horizontal=True, disabled=not use_resizer)
        if color_type == "Custom Color": 
            final_color_rgb = hex_to_rgb(st.color_picker("Pick color", "#FFFFFF"))
        elif color_type == "Standard White": 
            final_color_rgb = (255, 255, 255)
        else: 
            final_color_rgb = "AUTO"
            
        target_w, target_h = 660, 900
        if use_resizer:
            with st.expander("Adjust Dimensions"):
                target_w = st.number_input("Width (px)", min_value=10, value=660)
                target_h = st.number_input("Height (px)", min_value=10, value=900)

# STEP 1B: MANDATORY AI DISCLAIMER SCREEN
ai_liability_accepted = False
if use_ai_content:
    with st.container(border=True):
        st.markdown("<h4 style='color: #d9534f;'>⚠️ Mandatory AI Liability & Accuracy Disclaimer</h4>", unsafe_allow_html=True)
        st.info(
            "**Please Note:** AI model outputs are generated algorithmically. Gemini AI might occasionally produce "
            "inaccurate or mismatched classifications regarding categories, subtypes, or structural attributes. "
            "All generated metrics must be manually checked by an administrator prior to production use."
        )
        ai_liability_accepted = st.checkbox(
            "I acknowledge that using this automated content generation is entirely at our own risk. "
            "Our operational team accepts full responsibility for confirming the factual accuracy of all final data.",
            value=False
        )

# STEP 2: TEMPLATE HANDLING (THE BRAIN FILE)
mapping_context_str = ""
selected_ai_fields = []
template_headers = []
sku_header_col = None
link_header_col = None
content_tab = None

with st.container(border=True):
    st.subheader("📋 Step 2: Template & Mapping Configuration (Row 8 Headers)")
    template_file = st.file_uploader("Upload Blueprint/Mapping Excel Sheet", type=["xlsx"])
    
    if template_file:
        xl = pd.ExcelFile(template_file)
        tab_names = xl.sheet_names
        
        t1, t2 = st.columns(2)
        with t1:
            mapping_tab = st.selectbox("Select the **Mapping Rules** tab:", tab_names, index=0 if len(tab_names) > 0 else 0)
        with t2:
            content_tab = st.selectbox("Select the **Content Target** tab:", tab_names, index=1 if len(tab_names) > 1 else 0)
            
        try:
            df_mapping = pd.read_excel(template_file, sheet_name=mapping_tab, skiprows=7)
            mapping_context_str = df_mapping.to_string(index=False, max_rows=20)
            st.success("✅ Mapping rules scanned successfully from Row 8.")
        except Exception as e:
            st.error(f"Error reading Mapping tab on Row 8: {e}")
            
        try:
            temp_wb = openpyxl.load_workbook(template_file, read_only=True)
            if content_tab in temp_wb.sheetnames:
                temp_ws = temp_wb[content_tab]
                template_headers = [cell.value for cell in temp_ws[8] if cell.value is not None]
                template_headers = [str(h).strip() for h in template_headers if not str(h).startswith("Unnamed:")]
                
                if template_headers:
                    m1, m2 = st.columns(2)
                    with m1:
                        sku_header_col = st.selectbox("Which column is the **SKU / PSKU**?", template_headers)
                    with m2:
                        link_header_col = st.selectbox("Which column should receive the **Resized Image Link**?", ["None"] + template_headers, disabled=not use_resizer)
                    
                    if use_ai_content:
                        available_ai_headers = [h for h in template_headers if h != sku_header_col and h != link_header_col]
                        selected_ai_fields = st.multiselect("🎯 **Select columns you want Gemini AI to generate:**", available_ai_headers)
                else:
                    st.warning("No valid column headers found on Row 8 of the Content tab.")
            temp_wb.close()
        except Exception as e:
            st.error(f"Could not parse Row 8 headers from Content tab: {e}")

st.divider()

if use_ai_content and not GEMINI_API_KEY and selected_ai_fields:
    st.warning("⚠️ You selected AI Content generation, but your GEMINI_API_KEY is missing from Secrets.")

# --- DATA TARGET INPUT PREPARATION ---
data_to_process = []
df_original = None

if input_mode == "Links (Excel/CSV Sheet)":
    uploaded_file = st.file_uploader("Upload Product Target Sheet", type=["csv", "xlsx"])
    if uploaded_file:
        df_original = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
        c1, c2 = st.columns(2)
        with c1: sku_col = st.selectbox("Select Target SKU Column", df_original.columns)
        with c2: url_cols = st.multiselect("Select Target URL Column(s)", [c for c in df_original.columns if c != sku_col])
        if url_cols:
            for idx, row in df_original.iterrows():
                for col in url_cols:
                    data_to_process.append({"sku": str(row[sku_col]), "content": row[col], "col_name": col, "row_idx": idx, "type": "url"})
else:
    uploaded_imgs = st.file_uploader("Upload Target Images", type=["jpg", "png", "webp"], accept_multiple_files=True)
    if uploaded_imgs:
        for img_file in uploaded_imgs:
            data_to_process.append({"sku": img_file.name.rsplit('.', 1)[0], "content": Image.open(img_file), "col_name": "file", "type": "file"})

# --- VALIDATE AND RUN ---
if st.button("🚀 Start Production Loop") and data_to_process:
    if not use_resizer and not use_ai_content:
        st.error("🚨 You must tick either the Resizer or Content generation to execute a loop.")
        st.stop()
        
    if use_ai_content and not ai_liability_accepted:
        st.error("🚨 Execution Blocked: You must read and check the mandatory AI Liability & Accuracy Disclaimer box to run content features.")
        st.stop()
        
    if output_mode == "Links (Excel Sheet)" and not template_file:
        st.error("🚨 Please upload a Template Blueprint Sheet first to map the output columns.")
        st.stop()
        
    pb = st.progress(0)
    st_txt = st.empty()
    total = len(data_to_process)
    
    if output_mode == "Links (Excel Sheet)":
        template_file.seek(0)
        wb = openpyxl.load_workbook(template_file)
        ws = wb[content_tab]
        
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=8, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx
                
        sku_col_index = header_map.get(sku_header_col)
        link_col_index = header_map.get(link_header_col) if link_header_col != "None" else None
        
        if not sku_col_index:
            st.error(f"Could not locate the SKU column '{sku_header_col}' on Row 8 of your template.")
            st.stop()
            
        for i, item in enumerate(data_to_process):
            st_txt.text(f"Processing Loop {i+1}/{total}: {item['sku']}")
            try:
                proc_img, res_link = None, None
                
                if use_resizer:
                    if item['type'] == "file":
                        proc_img = process_image_pipeline(item['content'], target_w, target_h, final_color_rgb, bg_mode)
                        buf = BytesIO(); proc_img.save(buf, format="JPEG", quality=90); buf.seek(0)
                        res_link = upload_to_imgbb(buf, item['sku'])
                    else:
                        proc_img, res_link = process_url_full(item['content'], item['sku'], target_w, target_h, final_color_rgb, bg_mode)
                else:
                    if item['type'] == "file":
                        proc_img = item['content']
                    else:
                        resp = requests.get(get_direct_url(item['content']), timeout=15)
                        proc_img = Image.open(BytesIO(resp.content))
                
                target_row = None
                for row_idx in range(9, ws.max_row + 1):
                    cell_val = str(ws.cell(row=row_idx, column=sku_col_index).value).strip()
                    if cell_val == str(item['sku']).strip():
                        target_row = row_idx
                        break
                        
                if not target_row:
                    target_row = ws.max_row + 1
                    ws.cell(row=target_row, column=sku_col_index, value=item['sku'])
                    
                if use_resizer and link_col_index:
                    ws.cell(row=target_row, column=link_col_index, value=res_link)
                
                if use_ai_content and selected_ai_fields and GEMINI_API_KEY and proc_img:
                    st_txt.text(f"🤖 Gemini analyzing product {i+1}/{total}: {item['sku']}")
                    ai_outputs = generate_dynamic_content(proc_img, selected_ai_fields, mapping_context_str)
                    
                    for field in selected_ai_fields:
                        field_col_idx = header_map.get(field)
                        if field_col_idx:
                            ws.cell(row=target_row, column=field_col_idx, value=ai_outputs.get(field, ""))
                            
            except Exception as e: 
                st.error(f"Error on {item['sku']}: {e}")
            pb.progress((i + 1) / total)

        st.success("✅ Automation completed successfully!")
        out_excel = BytesIO()
        wb.save(out_excel)
        st.download_button("📥 Download Populated Workbook", out_excel.getvalue(), "Completed_Catalog_Template.xlsx")

    else:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_f:
            for i, item in enumerate(data_to_process):
                st_txt.text(f"Zipping {i+1}/{total}: {item['sku']}")
                try:
                    img = item['content'] if item['type'] == "file" else Image.open(BytesIO(requests.get(get_direct_url(item['content'])).content))
                    
                    if use_resizer:
                        proc_img = process_image_pipeline(img, target_w, target_h, final_color_rgb, bg_mode)
                        img_buf = BytesIO(); proc_img.save(img_buf, format="JPEG", quality=90)
                        zip_f.writestr(f"{item['sku']}.jpg", img_buf.getvalue())
                    else:
                        proc_img = img
                    
                    if use_ai_content and selected_ai_fields and GEMINI_API_KEY:
                        ai_outputs = generate_dynamic_content(proc_img, selected_ai_fields, mapping_context_str)
                        zip_f.writestr(f"{item['sku']}_metadata.json", json.dumps(ai_outputs, indent=4))
                except: pass
                pb.progress((i + 1) / total)
        st.success("✅ ZIP Generated!")
        st.download_button("📥 Download ZIP Package", zip_buffer.getvalue(), "processed_images.zip")
